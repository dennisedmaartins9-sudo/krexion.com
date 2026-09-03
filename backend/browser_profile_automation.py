"""
Browser Profile JSON automation (v2.7.79)
=========================================
Reuses RUT Visual Recorder step engine on headed profile sessions.
Default launch = manual browse. Automation runs ONLY when the user
explicitly enables JSON in the launch request.

v2.7.105c — Atomic lead-row claim (bulk race fix), consume_mode,
after_mode helpers, GSheet fingerprint consume, JSON↔data validate.
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("browser_profile_automation")

_ROW_CLAIM_LOCK_TTL_SEC = 45
_ROW_RESERVE_TTL_SEC = 3600
_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_.-]+)\s*\}\}|\{([a-zA-Z0-9_.-]+)\}")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _row_fingerprint(row: Dict[str, Any]) -> str:
    """Stable key for a lead row when email is missing (GSheet / audit)."""
    parts: List[str] = []
    for k in sorted((row or {}).keys()):
        if str(k).startswith("_"):
            continue
        v = row.get(k)
        if v is None or str(v).strip() == "":
            continue
        parts.append(f"{k}={str(v).strip()}")
    raw = "|".join(parts) or "empty"
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()[:24]


def _row_email(row: Dict[str, Any]) -> str:
    for k in ("email", "email_address", "emailaddress", "e_mail", "mail"):
        v = (row or {}).get(k)
        if v and str(v).strip():
            return str(v).strip()
    return ""


def _normalize_after_mode(raw: Any) -> str:
    m = str(raw or "manual").strip().lower()
    if m in ("close", "close_browser", "stop", "quit"):
        return "close"
    if m in ("next", "next_lead", "next_row", "loop"):
        return "next_lead"
    return "manual"


def _normalize_consume_mode(raw: Any) -> str:
    m = str(raw or "on_submit").strip().lower()
    if m in ("on_start", "start", "claim", "immediate"):
        return "on_start"
    return "on_submit"


def parse_automation_steps(raw: Any) -> List[Dict[str, Any]]:
    """Normalize upload payload → list of step dicts."""
    if isinstance(raw, list):
        return [dict(x) for x in raw if isinstance(x, dict)]
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
        except Exception as exc:
            raise ValueError(f"automation JSON parse failed: {exc}") from exc
        if isinstance(parsed, list):
            return [dict(x) for x in parsed if isinstance(x, dict)]
        if isinstance(parsed, dict) and isinstance(parsed.get("steps"), list):
            return [dict(x) for x in parsed["steps"] if isinstance(x, dict)]
        raise ValueError("automation JSON must be a list of steps")
    return []


async def load_automation_upload(
    db: Any,
    user_id: str,
    upload_id: str,
) -> Tuple[List[Dict[str, Any]], str]:
    """Load steps from an automation_json upload."""
    uid = str(upload_id or "").strip()
    if not uid:
        return [], ""
    doc = await db.uploaded_resources.find_one(
        {"id": uid, "user_id": user_id, "type": "automation_json"},
        {"_id": 0},
    )
    if not doc:
        raise ValueError("Automation JSON template not found")
    txt = str(doc.get("automation_json") or "").strip()
    items = doc.get("items") or []
    if not txt and items and isinstance(items, list):
        if len(items) == 1:
            txt = str(items[0])
        else:
            try:
                txt = json.dumps(items)
            except Exception:
                txt = ""
    steps = parse_automation_steps(txt)
    if not steps:
        raise ValueError("Automation template has no steps")
    name = str(doc.get("name") or doc.get("filename") or uid[:8])
    return steps, name


async def load_data_file_rows(db: Any, user_id: str, upload_id: str) -> List[Dict[str, Any]]:
    """Load lead rows from a data_file upload (same storage as RUT)."""
    uid = str(upload_id or "").strip()
    if not uid:
        return []
    doc = await db.uploaded_resources.find_one(
        {"id": uid, "user_id": user_id, "type": "data_file"},
        {"_id": 0},
    )
    if not doc:
        raise ValueError("Lead data file not found")
    remaining_fp = str(doc.get("remaining_file_path") or "").strip()
    file_path = (
        remaining_fp
        if remaining_fp and Path(remaining_fp).exists()
        else doc.get("file_path")
    )
    if not file_path or not Path(str(file_path)).exists():
        raise ValueError("Lead data file no longer on disk")
    from form_filler import load_rows_from_excel

    content = Path(str(file_path)).read_bytes()
    rows = load_rows_from_excel(content, doc.get("filename") or "data.xlsx")
    return [dict(r) for r in (rows or []) if isinstance(r, dict)]


async def _acquire_data_file_lock(db: Any, user_id: str, data_file_id: str) -> bool:
    """Short Mongo lock so bulk launches cannot pick the same lead row."""
    uid = str(data_file_id or "").strip()
    if not uid or db is None:
        return False
    now = datetime.now(timezone.utc)
    lock_until = (now + timedelta(seconds=_ROW_CLAIM_LOCK_TTL_SEC)).isoformat()
    stale_before = (now - timedelta(seconds=1)).isoformat()
    for _ in range(40):
        doc = await db.uploaded_resources.find_one_and_update(
            {
                "id": uid,
                "user_id": user_id,
                "type": "data_file",
                "$or": [
                    {"row_claim_lock_until": {"$exists": False}},
                    {"row_claim_lock_until": None},
                    {"row_claim_lock_until": ""},
                    {"row_claim_lock_until": {"$lte": stale_before}},
                ],
            },
            {"$set": {
                "row_claim_lock_until": lock_until,
                "row_claim_lock_at": now.isoformat(),
            }},
            return_document=True,
        )
        if doc:
            return True
        await asyncio.sleep(0.05)
    return False


async def _release_data_file_lock(db: Any, user_id: str, data_file_id: str) -> None:
    uid = str(data_file_id or "").strip()
    if not uid or db is None:
        return
    try:
        await db.uploaded_resources.update_one(
            {"id": uid, "user_id": user_id, "type": "data_file"},
            {"$set": {"row_claim_lock_until": "", "updated_at": _now_iso()}},
        )
    except Exception:
        pass


def _prune_reservations(doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw = doc.get("reserved_rows") or []
    if not isinstance(raw, list):
        return []
    now = datetime.now(timezone.utc)
    kept: List[Dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        exp = str(item.get("expires_at") or "")
        try:
            exp_dt = datetime.fromisoformat(exp.replace("Z", "+00:00"))
            if exp_dt.tzinfo is None:
                exp_dt = exp_dt.replace(tzinfo=timezone.utc)
            if exp_dt < now:
                continue
        except Exception:
            continue
        kept.append(item)
    return kept


async def claim_next_data_file_row(
    db: Any,
    user_id: str,
    data_file_id: str,
    *,
    consume_now: bool = False,
    prefer_index: Optional[int] = None,
    prefer_strict: bool = False,
) -> Tuple[Dict[str, Any], int, Dict[str, Any]]:
    """
    Atomically claim one lead row for a profile launch.

    Returns (lead_row, lead_row_index, meta).
    meta may include depleted / reserved / consumed_now.
    Raises ValueError when the file is empty or missing.
    """
    uid = str(data_file_id or "").strip()
    if not uid:
        raise ValueError("data_file_id required")
    if db is None:
        raise ValueError("database required for lead claim")

    locked = await _acquire_data_file_lock(db, user_id, uid)
    if not locked:
        raise ValueError("Lead file is busy — retry in a moment")

    meta: Dict[str, Any] = {"consumed_now": False, "reserved": False}
    try:
        doc = await db.uploaded_resources.find_one(
            {"id": uid, "user_id": user_id, "type": "data_file"},
            {"_id": 0},
        )
        if not doc:
            raise ValueError("Lead data file not found")
        if doc.get("depleted"):
            raise ValueError("Lead data file is depleted (no remaining rows)")

        rows = await load_data_file_rows(db, user_id, uid)
        if not rows:
            raise ValueError("Lead data file has no remaining rows")

        reserved = _prune_reservations(doc)
        reserved_idxs = {
            int(r.get("lead_row_index"))
            for r in reserved
            if r.get("lead_row_index") is not None
        }
        reserved_fps = {str(r.get("fingerprint") or "") for r in reserved if r.get("fingerprint")}
        consumed_keys = {
            str(x).strip().lower()
            for x in (doc.get("consumed_keys") or [])
            if str(x).strip()
        }

        def _available(i: int, row: Dict[str, Any]) -> bool:
            if i in reserved_idxs:
                return False
            fp = _row_fingerprint(row)
            if fp in reserved_fps:
                return False
            email = _row_email(row).lower()
            if email and email in consumed_keys:
                return False
            if fp.lower() in consumed_keys:
                return False
            return True

        chosen_idx: Optional[int] = None
        chosen_row: Dict[str, Any] = {}

        if prefer_index is not None:
            pi = int(prefer_index) % max(1, len(rows))
            if 0 <= pi < len(rows) and _available(pi, rows[pi]):
                chosen_idx = pi
                chosen_row = dict(rows[pi])
            elif prefer_strict:
                raise ValueError(
                    f"Lead row index {pi} is reserved, consumed, or unavailable"
                )

        if chosen_idx is None:
            for i, row in enumerate(rows):
                if _available(i, dict(row)):
                    chosen_idx = i
                    chosen_row = dict(row)
                    break

        if chosen_idx is None:
            raise ValueError("No free lead rows (all reserved or consumed)")

        fp = _row_fingerprint(chosen_row)
        if consume_now:
            ok = await consume_data_file_row(
                db, user_id, uid, int(chosen_idx), lead_row=chosen_row,
            )
            meta["consumed_now"] = bool(ok)
            if not ok:
                logger.warning(
                    f"[profile-automation] claim consume_now failed for row {chosen_idx}"
                )
        else:
            expires = (
                datetime.now(timezone.utc) + timedelta(seconds=_ROW_RESERVE_TTL_SEC)
            ).isoformat()
            reserved.append({
                "lead_row_index": int(chosen_idx),
                "fingerprint": fp,
                "email": _row_email(chosen_row)[:120],
                "expires_at": expires,
                "reserved_at": _now_iso(),
            })
            await db.uploaded_resources.update_one(
                {"id": uid, "user_id": user_id},
                {"$set": {
                    "reserved_rows": reserved[-200:],
                    "updated_at": _now_iso(),
                }},
            )
            meta["reserved"] = True

        meta["fingerprint"] = fp
        meta["remaining_after_pick"] = max(0, len(rows) - 1) if consume_now else max(0, len(rows) - len(reserved))
        return chosen_row, int(chosen_idx), meta
    finally:
        await _release_data_file_lock(db, user_id, uid)


async def release_reserved_data_file_row(
    db: Any,
    user_id: str,
    data_file_id: str,
    lead_row_index: Optional[int] = None,
    fingerprint: str = "",
) -> None:
    """Drop a soft reservation when automation is cancelled without consume."""
    uid = str(data_file_id or "").strip()
    if not uid or db is None:
        return
    doc = await db.uploaded_resources.find_one(
        {"id": uid, "user_id": user_id, "type": "data_file"},
        {"_id": 0, "reserved_rows": 1},
    )
    if not doc:
        return
    kept = []
    for item in _prune_reservations(doc):
        if lead_row_index is not None and int(item.get("lead_row_index") or -1) == int(lead_row_index):
            continue
        if fingerprint and str(item.get("fingerprint") or "") == str(fingerprint):
            continue
        kept.append(item)
    await db.uploaded_resources.update_one(
        {"id": uid, "user_id": user_id},
        {"$set": {"reserved_rows": kept, "updated_at": _now_iso()}},
    )


def extract_placeholders_from_steps(steps: List[Dict[str, Any]]) -> List[str]:
    found: set = set()
    for step in steps or []:
        if not isinstance(step, dict):
            continue
        for val in step.values():
            if isinstance(val, str):
                for m in _PLACEHOLDER_RE.finditer(val):
                    found.add((m.group(1) or m.group(2) or "").strip())
            elif isinstance(val, dict):
                for vv in val.values():
                    if isinstance(vv, str):
                        for m in _PLACEHOLDER_RE.finditer(vv):
                            found.add((m.group(1) or m.group(2) or "").strip())
    return sorted(x for x in found if x)


async def validate_automation_against_data(
    db: Any,
    user_id: str,
    *,
    automation_upload_id: str = "",
    data_file_id: str = "",
    steps: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Dry-run check: template parses + placeholders vs lead columns."""
    out: Dict[str, Any] = {
        "ok": True,
        "errors": [],
        "warnings": [],
        "total_steps": 0,
        "placeholders": [],
        "data_columns": [],
        "missing_columns": [],
        "row_count": 0,
        "depleted": False,
        "upload_name": "",
    }
    step_list: List[Dict[str, Any]] = list(steps or [])
    if automation_upload_id:
        try:
            step_list, name = await load_automation_upload(db, user_id, automation_upload_id)
            out["upload_name"] = name
        except Exception as exc:
            out["ok"] = False
            out["errors"].append(str(exc)[:240])
            return out
    out["total_steps"] = len(step_list)
    if not step_list:
        out["warnings"].append("Template has no steps")
    placeholders = extract_placeholders_from_steps(step_list)
    out["placeholders"] = placeholders

    dfid = str(data_file_id or "").strip()
    if dfid:
        try:
            doc = await db.uploaded_resources.find_one(
                {"id": dfid, "user_id": user_id, "type": "data_file"},
                {"_id": 0},
            )
            if not doc:
                out["ok"] = False
                out["errors"].append("Lead data file not found")
                return out
            out["depleted"] = bool(doc.get("depleted"))
            rows = await load_data_file_rows(db, user_id, dfid)
            out["row_count"] = len(rows)
            if not rows:
                out["ok"] = False
                out["errors"].append("Lead data file has no remaining rows")
            else:
                cols = sorted({str(k) for k in rows[0].keys() if not str(k).startswith("_")})
                out["data_columns"] = cols
                missing = [p for p in placeholders if p not in cols and p.lower() not in {c.lower() for c in cols}]
                out["missing_columns"] = missing
                if missing:
                    out["warnings"].append(
                        "Placeholders missing from lead columns: " + ", ".join(missing[:12])
                    )
        except Exception as exc:
            out["ok"] = False
            out["errors"].append(str(exc)[:240])
    elif placeholders:
        out["warnings"].append(
            "Template uses placeholders but no lead file selected: "
            + ", ".join(placeholders[:8])
        )
    if out["errors"]:
        out["ok"] = False
    return out


async def resolve_launch_automation(
    db: Any,
    user_id: str,
    automation: Optional[Dict[str, Any]],
    profile_doc: Optional[Dict[str, Any]] = None,
    *,
    bulk_row_index: Optional[int] = None,
    claim_next: Optional[bool] = None,
) -> Dict[str, Any]:
    """
    Build launch-time automation spec. Returns dict with:
      enabled, steps, lead_row, upload_name, total_steps, ...
    """
    prof = dict(profile_doc or {})
    auto = dict(automation or {})
    enabled = bool(auto.get("enabled"))

    upload_id = str(auto.get("automation_upload_id") or "").strip()
    data_file_id = str(auto.get("data_file_id") or "").strip()
    if not data_file_id:
        data_file_id = str(prof.get("default_data_file_id") or "").strip()

    consume_mode = _normalize_consume_mode(auto.get("consume_mode"))
    after_mode = _normalize_after_mode(auto.get("after_mode"))
    # Bulk always claims next free row (fixes index-collision race).
    use_claim = bool(claim_next) if claim_next is not None else (
        bulk_row_index is not None or bool(auto.get("claim_next"))
    )

    out: Dict[str, Any] = {
        "enabled": False,
        "steps": [],
        "lead_row": {},
        "automation_upload_id": upload_id,
        "data_file_id": data_file_id,
        "upload_name": "",
        "total_steps": 0,
        "lead_row_index": int(auto.get("lead_row_index") or 0),
        "skip_missing_steps": auto.get("skip_missing_steps", True) is not False,
        "self_heal": bool(auto.get("self_heal")),
        "after_mode": after_mode,
        "consume_mode": consume_mode,
        "claim_next": use_claim,
        "lead_consumed_already": False,
        "lead_fingerprint": "",
        "save_as_default": bool(auto.get("save_as_default")),
        "proxy_upload_id": str(auto.get("proxy_upload_id") or "").strip(),
        "ua_upload_id": str(auto.get("ua_upload_id") or "").strip(),
        "smart_funnel_enabled": bool(auto.get("smart_funnel_enabled")),
        "smart_funnel_pattern": str(auto.get("smart_funnel_pattern") or "auto"),
        "smart_funnel_min_deals": int(auto.get("smart_funnel_min_deals") or 2),
        "smart_funnel_wait_until_conversion": bool(auto.get("smart_funnel_wait_until_conversion", True)),
        "data_file_depleted": False,
        "data_file_warning": "",
    }

    if not enabled:
        return out

    if upload_id:
        steps, name = await load_automation_upload(db, user_id, upload_id)
        out["steps"] = steps
        out["upload_name"] = name
        out["total_steps"] = len(steps)
        out["enabled"] = True
    elif bool(auto.get("smart_funnel_enabled")):
        out["enabled"] = True
        out["total_steps"] = 1
    else:
        return out

    if data_file_id:
        prefer = None
        if not use_claim:
            prefer = int(auto.get("lead_row_index") or 0)
            if prefer < 0:
                prefer = 0
        elif bulk_row_index is not None and not bool(auto.get("claim_next", True)):
            # Legacy path only when caller explicitly disables claim_next
            prefer = int(bulk_row_index)

        try:
            if use_claim or prefer is not None:
                lead, idx, meta = await claim_next_data_file_row(
                    db,
                    user_id,
                    data_file_id,
                    consume_now=(consume_mode == "on_start"),
                    prefer_index=prefer if not use_claim else None,
                    prefer_strict=bool(prefer is not None and not use_claim),
                )
                out["lead_row"] = lead
                out["lead_row_index"] = idx
                out["lead_fingerprint"] = str(meta.get("fingerprint") or "")
                out["lead_consumed_already"] = bool(meta.get("consumed_now"))
            else:
                rows = await load_data_file_rows(db, user_id, data_file_id)
                if not rows:
                    out["data_file_depleted"] = True
                    out["data_file_warning"] = "Lead data file has no remaining rows"
                else:
                    idx = int(auto.get("lead_row_index") or 0)
                    if idx < 0:
                        idx = 0
                    idx = idx % len(rows)
                    out["lead_row"] = dict(rows[idx])
                    out["lead_row_index"] = idx
        except ValueError as exc:
            msg = str(exc)
            out["data_file_warning"] = msg
            if "depleted" in msg.lower() or "no remaining" in msg.lower() or "no free" in msg.lower():
                out["data_file_depleted"] = True
            # Fail hard when automation needs leads
            raise

    return out


async def consume_data_file_row(
    db: Any,
    user_id: str,
    data_file_id: str,
    row_index: int,
    *,
    lead_row: Optional[Dict[str, Any]] = None,
) -> bool:
    """Remove one consumed lead row from Uploaded Things (RUT parity)."""
    uid = str(data_file_id or "").strip()
    if not uid or db is None:
        return False
    doc = await db.uploaded_resources.find_one(
        {"id": uid, "user_id": user_id, "type": "data_file"},
        {"_id": 0},
    )
    if not doc:
        return False

    from datetime import datetime, timezone

    gsheet_url = str(doc.get("gsheet_url") or "").strip()
    if gsheet_url:
        rows = await load_data_file_rows(db, user_id, uid)
        target = dict(lead_row or {})
        if not target:
            if not rows or row_index < 0 or row_index >= len(rows):
                return False
            target = dict(rows[row_index] or {})
        target_email = _row_email(target)
        fp = _row_fingerprint(target)
        deleted = False
        try:
            import gsheet_writer

            loop = asyncio.get_running_loop()
            if target_email:
                deleted = await loop.run_in_executor(
                    None,
                    gsheet_writer.delete_row_by_email,
                    gsheet_url,
                    target_email,
                )
            if not deleted:
                # No email column / match — try first-column unique value
                first_val = ""
                for k, v in target.items():
                    if str(k).startswith("_"):
                        continue
                    if v is not None and str(v).strip():
                        first_val = str(v).strip()
                        break
                if first_val and hasattr(gsheet_writer, "delete_rows_by_first_column"):
                    n = await loop.run_in_executor(
                        None,
                        gsheet_writer.delete_rows_by_first_column,
                        gsheet_url,
                        [first_val],
                    )
                    deleted = bool(n)
        except Exception as exc:
            logger.debug(f"[profile-automation] gsheet consume skipped: {exc}")
        key_for_audit = (target_email or fp or f"row_{row_index}").lower()
        # Clear reservation + mark consumed even if remote delete failed
        reserved = [
            r for r in _prune_reservations(doc)
            if int(r.get("lead_row_index") or -1) != int(row_index)
            and str(r.get("fingerprint") or "") != fp
        ]
        await db.uploaded_resources.update_one(
            {"id": uid, "user_id": user_id},
            {
                "$addToSet": {"consumed_keys": key_for_audit},
                "$inc": {"consumed_count": 1, "item_count": -1},
                "$set": {
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "reserved_rows": reserved,
                },
            },
        )
        return bool(deleted or target_email or fp)

    try:
        import openpyxl
        import shutil

        UPLOADS_DATA_DIR = Path(__file__).resolve().parent / "uploaded_resources"
        UPLOADS_DATA_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        return False

    remaining_fp = str(doc.get("remaining_file_path") or "").strip()
    original_fp = str(doc.get("file_path") or "").strip()
    if not remaining_fp or not Path(remaining_fp).exists():
        if original_fp and Path(original_fp).exists():
            user_dir = UPLOADS_DATA_DIR / str(user_id)
            user_dir.mkdir(parents=True, exist_ok=True)
            remaining_fp = str(user_dir / f"{uid}__remaining.xlsx")
            shutil.copyfile(original_fp, remaining_fp)
            await db.uploaded_resources.update_one(
                {"id": uid, "user_id": user_id},
                {"$set": {"remaining_file_path": remaining_fp}},
            )
        else:
            return False

    fp_str = remaining_fp
    if not Path(fp_str).exists():
        return False

    try:
        wb = openpyxl.load_workbook(fp_str)
        ws = wb.active
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        display_header = [h for h in header if h and str(h) != "_orig_idx"]
        if "_orig_idx" not in header:
            col_idx = len(header) + 1
            ws.cell(row=1, column=col_idx, value="_orig_idx")
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx, value=r - 2)
            header.append("_orig_idx")
        orig_col = header.index("_orig_idx") + 1
        target_sheet_row = None
        row_values: List[Any] = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=orig_col).value
            try:
                if int(val) == int(row_index):
                    target_sheet_row = r
                    for c in range(1, len(header) + 1):
                        if header[c - 1] == "_orig_idx":
                            continue
                        row_values.append(ws.cell(row=r, column=c).value)
                    break
            except (TypeError, ValueError):
                continue
        if not target_sheet_row:
            wb.close()
            return False

        used_fp = str(doc.get("used_file_path") or "").strip()
        if not used_fp:
            user_dir = UPLOADS_DATA_DIR / str(user_id)
            user_dir.mkdir(parents=True, exist_ok=True)
            used_fp = str(user_dir / f"{uid}__used.xlsx")
        used_wb = openpyxl.load_workbook(used_fp) if Path(used_fp).exists() else openpyxl.Workbook()
        if "used data" in used_wb.sheetnames:
            used_ws = used_wb["used data"]
        else:
            used_ws = used_wb.active
            used_ws.title = "used data"
        if used_ws.max_row == 1 and not used_ws.cell(row=1, column=1).value:
            used_ws.delete_rows(1, 1)
        if used_ws.max_row < 1 or not used_ws.cell(row=1, column=1).value:
            for ci, h in enumerate(display_header or row_values, start=1):
                used_ws.cell(row=1, column=ci, value=h)
        next_used_row = used_ws.max_row + 1
        for ci, val in enumerate(row_values, start=1):
            used_ws.cell(row=next_used_row, column=ci, value=val)
        used_wb.save(used_fp)
        used_wb.close()

        ws.delete_rows(target_sheet_row, 1)
        wb.save(fp_str)
        remaining = max(0, (ws.max_row or 1) - 1)
        wb.close()

        fp = _row_fingerprint(dict(lead_row or {}))
        reserved = [
            r for r in _prune_reservations(doc)
            if int(r.get("lead_row_index") or -1) != int(row_index)
            and str(r.get("fingerprint") or "") != fp
        ]
        update: Dict[str, Any] = {
            "row_count": remaining,
            "item_count": remaining,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "remaining_file_path": fp_str,
            "used_file_path": used_fp,
            "reserved_rows": reserved,
        }
        await db.uploaded_resources.update_one(
            {"id": uid, "user_id": user_id},
            {"$set": update, "$inc": {"consumed_count": 1}},
        )
        if remaining == 0:
            try:
                Path(fp_str).unlink(missing_ok=True)
            except Exception:
                pass
            await db.uploaded_resources.update_one(
                {"id": uid, "user_id": user_id},
                {"$set": {
                    "depleted": True,
                    "depleted_at": datetime.now(timezone.utc).isoformat(),
                    "remaining_file_path": None,
                    "item_count": 0,
                    "row_count": 0,
                }},
            )
        return True
    except Exception as exc:
        logger.warning(f"[profile-automation] consume row failed: {exc}")
        return False


async def run_profile_automation(
    page: Any,
    steps: List[Dict[str, Any]],
    lead_row: Dict[str, Any],
    *,
    user_id: str,
    session_id: str,
    profile_id: str,
    on_session_update: Optional[Any] = None,
    skip_missing_steps: bool = True,
    self_heal: bool = False,
    should_cancel: Optional[Any] = None,
    data_file_id: str = "",
    lead_row_index: Optional[int] = None,
    db: Any = None,
    on_lead_submitted: Optional[Any] = None,
    smart_funnel_enabled: bool = False,
    smart_funnel_pattern: str = "auto",
    smart_funnel_min_deals: int = 2,
    smart_funnel_wait_until_conversion: bool = True,
    consume_mode: str = "on_submit",
    lead_consumed_already: bool = False,
    lead_fingerprint: str = "",
    after_mode: str = "manual",
) -> Dict[str, Any]:
    """Execute RUT automation steps on an open profile page."""
    from real_user_traffic import _execute_automation_steps

    total = len(steps or [])
    if not steps and not smart_funnel_enabled:
        return {"status": "skipped", "executed_steps": 0, "total_steps": 0, "after_mode": _normalize_after_mode(after_mode)}

    _lead_consumed = bool(lead_consumed_already)
    _consume_mode = _normalize_consume_mode(consume_mode)
    _after_mode = _normalize_after_mode(after_mode)
    _fp = str(lead_fingerprint or "") or _row_fingerprint(dict(lead_row or {}))

    async def _consume_lead(reason: str) -> None:
        nonlocal _lead_consumed
        if _lead_consumed:
            return
        if on_lead_submitted is not None:
            _lead_consumed = True
            await on_lead_submitted(str(reason or "form_submit")[:120])
            return
        if db is None or not data_file_id or lead_row_index is None:
            return
        _lead_consumed = True
        try:
            await consume_data_file_row(
                db, user_id, data_file_id, int(lead_row_index), lead_row=dict(lead_row or {}),
            )
            try:
                from browser_profile_audit import log_profile_event

                await log_profile_event(
                    db,
                    user_id=user_id,
                    profile_id=profile_id,
                    session_id=session_id,
                    event_type="lead_consumed",
                    detail={
                        "data_file_id": data_file_id,
                        "lead_row_index": int(lead_row_index),
                        "reason": str(reason or "")[:120],
                        "lead_email": _row_email(lead_row or {})[:120],
                        "fingerprint": _fp[:32],
                    },
                )
            except Exception:
                pass
            logger.info(
                f"[profile-automation] lead row #{int(lead_row_index) + 1} consumed ({reason})"
            )
        except Exception as exc:
            _lead_consumed = False
            logger.debug(f"[profile-automation] lead consume failed: {exc}")

    _lead_cb = _consume_lead if (data_file_id and lead_row_index is not None and not _lead_consumed) else on_lead_submitted

    # Optional: consume at automation start (already claimed+consumed when lead_consumed_already)
    if (
        not _lead_consumed
        and _consume_mode == "on_start"
        and data_file_id
        and lead_row_index is not None
    ):
        await _consume_lead("on_start")

    async def _on_progress(payload: Dict[str, Any]) -> None:
        if on_session_update is None:
            return
        try:
            idx = int(payload.get("idx") or payload.get("step_index") or 0)
            await on_session_update({
                "profile_id": profile_id,
                "session_id": session_id,
                "status": "running",
                "automation_status": "running",
                "automation_step": idx + 1,
                "automation_total_steps": total or 1,
                "automation_action": str(payload.get("action") or "")[:64],
            })
        except Exception as exc:
            logger.debug(f"[profile-automation] progress callback skipped: {exc}")

    if smart_funnel_enabled:
        try:
            from smart_funnel import execute_smart_funnel, rut_config_for_pattern

            def _substitute(tmpl: str, row: Dict[str, Any]) -> str:
                out = str(tmpl or "")
                for k, v in (row or {}).items():
                    out = out.replace("{{" + str(k) + "}}", str(v or ""))
                    out = out.replace("{" + str(k) + "}", str(v or ""))
                return out

            sf_cfg = rut_config_for_pattern(
                str(smart_funnel_pattern or "auto"),
                min_deals=max(1, min(5, int(smart_funnel_min_deals or 2))),
                wait_until_conversion=bool(smart_funnel_wait_until_conversion),
            )
            if on_session_update:
                await on_session_update({
                    "profile_id": profile_id,
                    "session_id": session_id,
                    "status": "running",
                    "automation_status": "running",
                    "automation_step": 0,
                    "automation_total_steps": 1,
                    "automation_action": "smart_funnel",
                })
            result = await execute_smart_funnel(
                page,
                dict(lead_row or {}),
                substitute=_substitute,
                config=sf_cfg,
                on_step_progress=_on_progress,
                job_id=f"profile-{session_id[:8]}",
                on_lead_submitted=_lead_cb,
            )
            auto_status = "completed" if result.get("status") == "ok" else str(result.get("status") or "failed")
            if on_session_update:
                await on_session_update({
                    "profile_id": profile_id,
                    "session_id": session_id,
                    "status": "running",
                    "automation_status": auto_status,
                    "automation_step": 1,
                    "automation_total_steps": 1,
                    "automation_error": str(result.get("error") or "")[:512],
                    "automation_lead_consumed": _lead_consumed,
                    "automation_after_mode": _after_mode,
                })
            if isinstance(result, dict):
                result["after_mode"] = _after_mode
                result["lead_consumed"] = _lead_consumed
            return result
        except Exception as exc:
            logger.warning(f"[profile-automation] smart funnel failed: {exc}")
            if on_session_update:
                try:
                    await on_session_update({
                        "profile_id": profile_id,
                        "session_id": session_id,
                        "status": "running",
                        "automation_status": "error",
                        "automation_error": str(exc)[:512],
                    })
                except Exception:
                    pass
            return {"status": "failed", "error": str(exc), "after_mode": _after_mode, "lead_consumed": _lead_consumed}

    try:
        if on_session_update:
            await on_session_update({
                "profile_id": profile_id,
                "session_id": session_id,
                "status": "running",
                "automation_status": "running",
                "automation_step": 0,
                "automation_total_steps": total,
            })
    except Exception:
        pass

    result = await _execute_automation_steps(
        page,
        dict(lead_row or {}),
        steps,
        skip_captcha=True,
        self_heal=bool(self_heal),
        user_id=str(user_id or "") or None,
        on_step_progress=_on_progress,
        skip_missing_steps=bool(skip_missing_steps),
        job_id=f"profile-{session_id[:8]}",
        should_cancel=should_cancel,
        on_lead_submitted=_lead_cb,
    )

    auto_status = str(result.get("status") or "failed")
    if auto_status == "cancelled":
        auto_status = "stopped"
    elif auto_status in ("ok", "skipped_captcha"):
        auto_status = "completed" if auto_status == "ok" else "captcha"

    if on_session_update:
        try:
            await on_session_update({
                "profile_id": profile_id,
                "session_id": session_id,
                "status": "running",
                "automation_status": auto_status,
                "automation_step": int(result.get("executed_steps") or 0),
                "automation_total_steps": total,
                "automation_error": str(result.get("error") or "")[:512],
                "automation_lead_consumed": _lead_consumed,
                "automation_after_mode": _after_mode,
            })
        except Exception:
            pass

    if isinstance(result, dict):
        result["after_mode"] = _after_mode
        result["lead_consumed"] = _lead_consumed
        result["automation_status"] = auto_status
    return result
